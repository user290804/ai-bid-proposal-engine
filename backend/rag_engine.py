import os
import openpyxl
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

class CapabilityRAG:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.capabilities = []
        self.vectorizer = None
        self.tfidf_matrix = None
        self.load_capabilities()
        self.index_capabilities()

    def load_capabilities(self):
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Dataset file not found at {self.excel_path}")

        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        # Sheetnames[1] is the Capability Library
        sheet = wb[wb.sheetnames[1]]
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Row 3 contains headers
        headers = [sheet.cell(row=3, column=c).value for c in range(1, max_col + 1)]

        for r in range(4, max_row + 1):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            if any(row_vals):
                self.capabilities.append(dict(zip(headers, row_vals)))
        
        print(f"Loaded {len(self.capabilities)} capabilities for RAG database.")

    def index_capabilities(self):
        corpus = []
        for cap in self.capabilities:
            domain = str(cap.get('Domain', '') or '')
            summary = str(cap.get('Project Summary', '') or '')
            cert = str(cap.get('Certification', '') or '')
            combined_text = f"{domain} {summary} {cert}"
            corpus.append(combined_text)

        self.vectorizer = TfidfVectorizer(stop_words='english')
        self.tfidf_matrix = self.vectorizer.fit_transform(corpus)
        print("Indexed capabilities successfully.")

    def match_requirement(self, req_text: str, threshold: float = 0.12):
        if not self.capabilities or not self.vectorizer:
            return {"status": "Fail", "score": 0.0, "evidence": None, "reason": "RAG database is empty"}

        # Transform query and compute cosine similarities
        query_vec = self.vectorizer.transform([req_text])
        similarities = cosine_similarity(query_vec, self.tfidf_matrix)[0]
        
        best_idx = int(similarities.argmax())
        best_score = float(similarities[best_idx])

        # Intelligent certification matching
        req_text_upper = req_text.upper()
        cert_required = None
        for cert in ["ISO 27001", "CMMI L3", "CE MARK", "ISO 9001", "PMP"]:
            if cert in req_text_upper:
                cert_required = cert
                break

        matched_cap = self.capabilities[best_idx]
        
        # If a specific certification is requested, verify if our candidate has it.
        # If not, look for any capability in our library that does have this certification.
        if cert_required:
            matched_cert = str(matched_cap.get('Certification', '') or '').upper()
            if cert_required not in matched_cert:
                cert_found = False
                for cap in self.capabilities:
                    if cert_required in str(cap.get('Certification', '') or '').upper():
                        matched_cap = cap
                        best_score = max(best_score, 0.45)  # boost similarity because certification matches
                        cert_found = True
                        break
                
                # If no project in the entire company has this certification, it's a hard compliance gap
                if not cert_found:
                    return {
                        "status": "Fail",
                        "score": round(best_score, 4),
                        "evidence": None,
                        "reason": f"Required certification '{cert_required}' is not possessed by the organization."
                    }

        # Apply threshold check
        if best_score >= threshold:
            return {
                "status": "Pass",
                "score": round(best_score, 4),
                "evidence": {
                    "cap_id": matched_cap.get('Cap ID'),
                    "domain": matched_cap.get('Domain'),
                    "summary": matched_cap.get('Project Summary'),
                    "certification": matched_cap.get('Certification'),
                    "year_completed": matched_cap.get('Year Completed'),
                    "contract_value": matched_cap.get('Contract Value'),
                    "client_type": matched_cap.get('Client Type')
                }
            }
        else:
            return {
                "status": "Fail",
                "score": round(best_score, 4),
                "evidence": None,
                "reason": "No matching capability matches the requirement description."
            }
