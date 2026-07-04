import os
import openpyxl
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import OneHotEncoder, StandardScaler
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline

class BidScoringModel:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.pipeline = None
        self.train_model()

    def parse_budget(self, val) -> float:
        if val is None:
            return 0.0
        val_str = str(val).upper().strip()
        is_m = "M" in val_str
        is_k = "K" in val_str
        # Clean string to get numeric value
        val_str = val_str.replace("PKR", "").replace("M", "").replace("K", "").replace(",", "").strip()
        try:
            num = float(val_str)
            if is_m:
                return num
            elif is_k:
                return num / 1000.0
            return num
        except ValueError:
            return 0.0

    def train_model(self):
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Dataset file not found at {self.excel_path}")

        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        # Access the first sheet for Bid History
        sheet = wb[wb.sheetnames[0]]
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Row 3 has headers
        headers = [sheet.cell(row=3, column=c).value for c in range(1, max_col + 1)]

        data = []
        for r in range(4, max_row + 1):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            if any(row_vals):
                data.append(dict(zip(headers, row_vals)))

        X = []
        y = []

        def safe_float(val) -> float:
            try:
                return float(val) if val is not None else 0.0
            except (ValueError, TypeError):
                return 0.0

        for row in data:
            if row.get('Outcome') is None or row.get('Sector') is None:
                continue

            # Outcome: Win -> 1, Loss -> 0
            outcome_val = str(row['Outcome']).strip().lower()
            y.append(1 if outcome_val == 'win' else 0)

            # Features
            X.append([
                str(row['Sector']).strip(),
                self.parse_budget(row.get('Budget')),
                safe_float(row.get('Compliance %')),
                safe_float(row.get('Doc Pages')),
                safe_float(row.get('Gaps Found')),
                safe_float(row.get('Response Time (hrs)'))
            ])

        X = np.array(X, dtype=object)
        y = np.array(y)

        # Model Pipeline definition
        categorical_features = [0]
        numeric_features = [1, 2, 3, 4, 5]

        preprocessor = ColumnTransformer(
            transformers=[
                ('cat', OneHotEncoder(handle_unknown='ignore'), categorical_features),
                ('num', StandardScaler(), numeric_features)
            ]
        )

        self.pipeline = Pipeline(steps=[
            ('preprocessor', preprocessor),
            ('classifier', RandomForestClassifier(n_estimators=100, random_state=42))
        ])

        self.pipeline.fit(X, y)
        print(f"Scoring Model successfully trained on {len(y)} historical records.")

    def predict_win_probability(self, sector: str, budget_m: float, compliance_pct: float, gaps_found: int, doc_pages: int, response_time_hrs: float) -> float:
        if self.pipeline is None:
            return 50.0
        
        input_data = np.array([[
            sector,
            budget_m,
            compliance_pct,
            doc_pages,
            gaps_found,
            response_time_hrs
        ]], dtype=object)

        probs = self.pipeline.predict_proba(input_data)[0]
        win_prob = float(probs[1]) * 100.0
        return round(win_prob, 2)
